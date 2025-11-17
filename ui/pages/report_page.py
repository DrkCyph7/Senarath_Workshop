"""
Advanced Analytics & Reporting Module
Provides comprehensive insights into workshop operations with export capabilities
"""
import sqlite3
import csv
import json
import datetime
from collections import defaultdict, Counter
from pathlib import Path
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QMessageBox, QComboBox, QDateEdit,
    QFrame, QFileDialog, QHeaderView, QTabWidget, QGridLayout,
    QDialog, QDialogButtonBox
)
from PySide6.QtCore import Qt, QDate
from PySide6.QtGui import QFont
from ui.theme import ColorPalette, Typography, Spacing, Styles, create_page_header

DB_PATH = "ui/db/senarath.db"

# Check for optional libraries
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

try:
    import openpyxl
    from openpyxl.styles import Font as XLFont, Alignment, PatternFill, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ReportPage(QWidget):
    """Advanced analytics dashboard with comprehensive reporting capabilities"""
    
    def __init__(self, parent=None):
        super().__init__()
        self.parent = parent
        self.current_report_data = None
        self.filtered_jobcards = []
        self._ensure_indexes()
        
        bg_color = "#f5f5f0"
        card_color = ColorPalette.CARD_BG
        accent_color = ColorPalette.ACCENT_PRIMARY
        text_primary = "#2c2c2c"
        text_secondary = ColorPalette.TEXT_SECONDARY
        border_color = "#d4d4d4"

        self.setStyleSheet(f"""
            QWidget {{
                background-color: {bg_color};
                font-family: 'Segoe UI', Arial;
                color: {text_primary};
            }}
            QLabel {{
                background-color: transparent;
            }}
            QFrame#stat_card {{
                background-color: {card_color};
                border: 1px solid {border_color};
                border-radius: {Spacing.BORDER_RADIUS_MEDIUM}px;
            }}
            QFrame#filter_card {{
                background-color: {card_color};
                border: 1px solid {border_color};
                border-radius: {Spacing.BORDER_RADIUS_SMALL}px;
            }}
            QLabel#stat_value {{
                color: {accent_color};
                font-size: 24px;
                font-weight: 700;
            }}
            QLabel#stat_label {{
                color: {text_secondary};
                font-size: 11px;
                font-weight: 600;
                text-transform: uppercase;
            }}
            QLabel#section_title {{
                color: {text_primary};
                font-size: 14px;
                font-weight: 700;
            }}
            QComboBox, QDateEdit {{
                background-color: #fafafa;
                border: 1px solid {border_color};
                border-radius: {Spacing.BORDER_RADIUS_SMALL}px;
                padding: 6px 8px;
                font-size: 12px;
                min-height: 24px;
                color: {text_primary};
            }}
            QComboBox:focus, QDateEdit:focus {{
                border: 2px solid {accent_color};
                background-color: white;
            }}
            QTableWidget {{
                background-color: {card_color};
                border: 1px solid {border_color};
                border-radius: {Spacing.BORDER_RADIUS_SMALL}px;
                gridline-color: {border_color};
                alternate-background-color: #f8f8f3;
                font-size: 11px;
            }}
            QHeaderView::section {{
                background-color: {accent_color};
                color: white;
                padding: 8px;
                border: none;
                font-weight: 600;
                font-size: 11px;
            }}
            QTableWidget::item {{
                padding: 4px;
                color: {text_primary};
            }}
            QTableWidget::item:selected {{
                background-color: rgba(45, 122, 95, 0.18);
            }}
            QTabWidget::pane {{
                border: 1px solid {border_color};
                border-radius: {Spacing.BORDER_RADIUS_SMALL}px;
                background-color: {card_color};
            }}
            QTabBar::tab {{
                background-color: transparent;
                padding: 8px 16px;
                font-weight: 600;
                color: {text_secondary};
                border: none;
            }}
            QTabBar::tab:selected {{
                color: {accent_color};
                border-bottom: 3px solid {accent_color};
            }}
        """)
        
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(18, 18, 18, 18)
        main_layout.setSpacing(16)
        
        # Header
        header_layout, title_label, back_btn = create_page_header("📊 Analytics & Reports")
        back_btn.clicked.connect(self.go_back)
        main_layout.addLayout(header_layout)
        
        # Filter Panel
        filter_frame = QFrame()
        filter_frame.setObjectName("filter_card")
        filter_layout = QVBoxLayout(filter_frame)
        filter_layout.setContentsMargins(16, 14, 16, 14)
        filter_layout.setSpacing(8)
        
        filter_title = QLabel("Report Filters")
        filter_title.setObjectName("section_title")
        filter_layout.addWidget(filter_title)
        
        # Filter controls laid out on compact grid
        controls_grid = QGridLayout()
        controls_grid.setContentsMargins(0, 0, 0, 0)
        controls_grid.setHorizontalSpacing(12)
        controls_grid.setVerticalSpacing(8)

        def _add_filter(label_text: str, widget, row: int, col: int):
            label = QLabel(label_text)
            label.setStyleSheet(f"color: {text_secondary}; font-size: 11px; font-weight: 600;")
            controls_grid.addWidget(label, row, col)
            controls_grid.addWidget(widget, row, col + 1)

        self.period_combo = QComboBox()
        self.period_combo.addItems(["Last 7 Days", "Last 30 Days", "Last 3 Months", "Last 6 Months", "This Year", "Custom Range"])
        self.period_combo.setMinimumWidth(140)
        self.period_combo.currentTextChanged.connect(self.on_period_changed)
        _add_filter("Period", self.period_combo, 0, 0)

        date_row = 0
        self.from_date = QDateEdit()
        self.from_date.setCalendarPopup(True)
        self.from_date.setDisplayFormat("yyyy-MM-dd")
        self.from_date.setDate(QDate.currentDate().addDays(-30))
        self.from_date.setEnabled(False)
        self.from_date.setMinimumWidth(130)
        _add_filter("From", self.from_date, date_row, 2)

        self.to_date = QDateEdit()
        self.to_date.setCalendarPopup(True)
        self.to_date.setDisplayFormat("yyyy-MM-dd")
        self.to_date.setDate(QDate.currentDate())
        self.to_date.setEnabled(False)
        self.to_date.setMinimumWidth(120)
        self.to_date.setMaximumWidth(160)
        to_label = QLabel("To")
        to_label.setStyleSheet(f"color: {text_secondary}; font-size: 11px; font-weight: 600;")
        controls_grid.addWidget(to_label, date_row, 4, alignment=Qt.AlignRight | Qt.AlignVCenter)
        controls_grid.addWidget(self.to_date, date_row, 5)

        self.site_filter = QComboBox()
        self.site_filter.addItem("All Sites")
        self.load_sites()
        self.site_filter.setMinimumWidth(150)
        _add_filter("Site", self.site_filter, 1, 0)

        self.status_filter = QComboBox()
        self.status_filter.addItems(["All Status", "Completed", "In Progress"])
        self.status_filter.setMinimumWidth(130)
        _add_filter("Status", self.status_filter, 1, 2)

        # Generate button aligned to the right column spanning rows
        generate_btn = QPushButton("Generate")
        generate_btn.setStyleSheet(Styles.get_button_primary())
        generate_btn.setFixedHeight(28)
        generate_btn.setCursor(Qt.PointingHandCursor)
        generate_btn.clicked.connect(self.generate_report)
        controls_grid.addWidget(generate_btn, 0, 6, 2, 1, alignment=Qt.AlignRight | Qt.AlignVCenter)

        # Spacer to keep grid compact but responsive
        controls_grid.setColumnStretch(5, 0)
        controls_grid.setColumnStretch(6, 1)

        filter_layout.addLayout(controls_grid)
        main_layout.addWidget(filter_frame)
        
        # Stats Cards
        self.stats_container = QHBoxLayout()
        self.stats_container.setSpacing(12)
        main_layout.addLayout(self.stats_container)
        
        # Tabs for different report views
        self.tabs = QTabWidget()
        self.tabs.addTab(self._create_overview_tab(), "📊 Overview")
        self.tabs.addTab(self._create_engine_oil_tab(), "🛢 Engine Oils")
        self.tabs.addTab(self._create_vehicle_insights_tab(), "🚗 Vehicle Insights")
        self.tabs.addTab(self._create_spare_part_trends_tab(), "🔧 Spare Part Trends")
        self.tabs.addTab(self._create_cost_trends_tab(), "💰 Cost Trends")
        main_layout.addWidget(self.tabs, 1)
        
        # Export buttons
        export_layout = QHBoxLayout()
        export_layout.addStretch()
        
        csv_btn = QPushButton("📄 Export CSV")
        csv_btn.setStyleSheet(Styles.get_button_secondary())
        csv_btn.setFixedHeight(28)
        csv_btn.setCursor(Qt.PointingHandCursor)
        csv_btn.clicked.connect(self.export_csv)
        export_layout.addWidget(csv_btn)
        
        if HAS_OPENPYXL:
            xlsx_btn = QPushButton("📊 Export Excel")
            xlsx_btn.setStyleSheet(Styles.get_button_secondary())
            xlsx_btn.setFixedHeight(28)
            xlsx_btn.setCursor(Qt.PointingHandCursor)
            xlsx_btn.clicked.connect(self.export_xlsx)
            export_layout.addWidget(xlsx_btn)
        
        if HAS_REPORTLAB:
            pdf_btn = QPushButton("📑 Export PDF")
            pdf_btn.setStyleSheet(Styles.get_button_secondary())
            pdf_btn.setFixedHeight(28)
            pdf_btn.setCursor(Qt.PointingHandCursor)
            pdf_btn.clicked.connect(self.export_pdf)
            export_layout.addWidget(pdf_btn)
        
        main_layout.addLayout(export_layout)
        
        # Initial data load
        self.generate_report()
    
    def _create_overview_tab(self):
        """Overview statistics and summary"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(12)
        
        self.overview_table = QTableWidget()
        self.overview_table.setAlternatingRowColors(True)
        self.overview_table.verticalHeader().setVisible(False)
        self.overview_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.overview_table.setSelectionMode(QTableWidget.NoSelection)
        layout.addWidget(self.overview_table)
        
        return tab
    
    def _create_engine_oil_tab(self):
        """Engine oil usage analysis"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(12)
        
        hint = QLabel("Double-click a row to drill into related job cards")
        hint.setStyleSheet("color: #6b7280; font-size: 11px; font-weight: 500;")
        layout.addWidget(hint)

        self.engine_oil_table = QTableWidget()
        self.engine_oil_table.setAlternatingRowColors(True)
        self.engine_oil_table.verticalHeader().setVisible(False)
        self.engine_oil_table.setEditTriggers(QTableWidget.NoEditTriggers)
        layout.addWidget(self.engine_oil_table)
        self._register_drilldown(self.engine_oil_table)
        
        return tab
    
    def _create_vehicle_insights_tab(self):
        """Vehicle servicing frequency and repeat repair analysis"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(12)
        
        vehicle_hint = QLabel("Double-click to open job history for the selected vehicle")
        vehicle_hint.setStyleSheet("color: #6b7280; font-size: 11px; font-weight: 500;")
        layout.addWidget(vehicle_hint)

        self.vehicle_table = QTableWidget()
        self.vehicle_table.setAlternatingRowColors(True)
        self.vehicle_table.verticalHeader().setVisible(False)
        self.vehicle_table.setEditTriggers(QTableWidget.NoEditTriggers)
        layout.addWidget(self.vehicle_table)
        self._register_drilldown(self.vehicle_table)

        repeat_label = QLabel("Repeat Repairs by Vehicle & Section")
        repeat_label.setStyleSheet("color: #374151; font-size: 12px; font-weight: 600;")
        layout.addWidget(repeat_label)

        self.repeat_repairs_table = QTableWidget()
        self.repeat_repairs_table.setAlternatingRowColors(True)
        self.repeat_repairs_table.verticalHeader().setVisible(False)
        self.repeat_repairs_table.setEditTriggers(QTableWidget.NoEditTriggers)
        layout.addWidget(self.repeat_repairs_table)
        self._register_drilldown(self.repeat_repairs_table)

        return tab

    def _create_spare_part_trends_tab(self):
        """Top spare parts within the selected window"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(12)

        hint = QLabel("Focuses on the trailing 7 days within the selected range; double-click for linked jobs")
        hint.setStyleSheet("color: #6b7280; font-size: 11px; font-weight: 500;")
        layout.addWidget(hint)

        self.spare_trends_table = QTableWidget()
        self.spare_trends_table.setAlternatingRowColors(True)
        self.spare_trends_table.verticalHeader().setVisible(False)
        self.spare_trends_table.setEditTriggers(QTableWidget.NoEditTriggers)
        layout.addWidget(self.spare_trends_table)
        self._register_drilldown(self.spare_trends_table)

        return tab

    def _create_cost_trends_tab(self):
        """Cost breakdown aggregated by month"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(12)

        self.cost_table = QTableWidget()
        self.cost_table.setAlternatingRowColors(True)
        self.cost_table.verticalHeader().setVisible(False)
        self.cost_table.setEditTriggers(QTableWidget.NoEditTriggers)
        layout.addWidget(self.cost_table)
        self._register_drilldown(self.cost_table)

        return tab
    
    def load_sites(self):
        """Load sites from database"""
        try:
            conn = sqlite3.connect(DB_PATH)
            c = conn.cursor()
            c.execute("SELECT DISTINCT name FROM sites ORDER BY name")
            sites = c.fetchall()
            conn.close()
            for site in sites:
                self.site_filter.addItem(site[0])
        except Exception as e:
            print(f"Error loading sites: {e}")
    
    def on_period_changed(self, period):
        """Handle period selection changes"""
        is_custom = period == "Custom Range"
        self.from_date.setEnabled(is_custom)
        self.to_date.setEnabled(is_custom)
        
        if not is_custom:
            today = QDate.currentDate()
            if period == "Last 7 Days":
                self.from_date.setDate(today.addDays(-7))
            elif period == "Last 30 Days":
                self.from_date.setDate(today.addDays(-30))
            elif period == "Last 3 Months":
                self.from_date.setDate(today.addMonths(-3))
            elif period == "Last 6 Months":
                self.from_date.setDate(today.addMonths(-6))
            elif period == "This Year":
                self.from_date.setDate(QDate(today.year(), 1, 1))
            self.to_date.setDate(today)
    
    def generate_report(self):
        """Generate comprehensive analytics report"""
        try:
            filters = self._build_filter_clause()
            self.filtered_jobcards = self._get_filtered_jobcards(filters)
            self._update_stats_cards(self.filtered_jobcards)
            self._update_overview_tab(self.filtered_jobcards)
            self._load_engine_oil_usage(filters)
            self._load_vehicle_insights(filters)
            self._load_spare_part_trends(filters)
            self._load_cost_trends(filters)
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to generate report:\n{str(e)}")
    
    def _update_stats_cards(self, job_cards):
        """Update top-level statistics cards"""
        # Clear existing cards
        while self.stats_container.count():
            item = self.stats_container.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        
        total_jobs = len(job_cards)
        completed_jobs = sum(1 for job in job_cards if job.get('status') == "Completed")
        in_progress = total_jobs - completed_jobs

        total_spare_cost = sum(job.get('spare_cost', 0.0) for job in job_cards)
        total_labour_cost = sum(job.get('labour_cost', 0.0) for job in job_cards)
        total_outsource_cost = sum(job.get('outsource_cost', 0.0) for job in job_cards)
        total_cost = total_spare_cost + total_labour_cost + total_outsource_cost

        self._add_stat_card("Total Jobs", str(total_jobs), ColorPalette.ACCENT_PRIMARY)
        self._add_stat_card("Completed", str(completed_jobs), ColorPalette.ACCENT_GREEN)
        self._add_stat_card("In Progress", str(in_progress), ColorPalette.ACCENT_ORANGE)
        self._add_stat_card("Total Cost", self._format_currency(total_cost), ColorPalette.ACCENT_SECONDARY)
    
    def _add_stat_card(self, label, value, color):
        """Add a statistics card"""
        card = QFrame()
        card.setObjectName("stat_card")
        layout = QVBoxLayout(card)
        layout.setContentsMargins(16, 12, 16, 12)
        layout.setSpacing(4)
        
        value_label = QLabel(value)
        value_label.setObjectName("stat_value")
        value_label.setStyleSheet(f"color: {color};")
        
        text_label = QLabel(label)
        text_label.setObjectName("stat_label")
        
        layout.addWidget(value_label)
        layout.addWidget(text_label)
        
        self.stats_container.addWidget(card)
    
    def _update_overview_tab(self, job_cards):
        """Update overview statistics"""
        headers = ["Metric", "Value"]
        self.overview_table.setColumnCount(2)
        self.overview_table.setHorizontalHeaderLabels(headers)
        
        total_jobs = len(job_cards)
        completed_jobs = sum(1 for job in job_cards if job.get('status') == "Completed")
        in_progress = total_jobs - completed_jobs

        unique_vehicles = {job.get('vehicle_no') for job in job_cards if job.get('vehicle_no') and job.get('vehicle_no') not in {"-", "", "N/A"}}
        unique_drivers = {job.get('driver') for job in job_cards if job.get('driver')}
        unique_sites = {job.get('site') for job in job_cards if job.get('site')}

        total_spare = sum(job.get('spare_cost', 0.0) for job in job_cards)
        total_labour = sum(job.get('labour_cost', 0.0) for job in job_cards)
        total_outsource = sum(job.get('outsource_cost', 0.0) for job in job_cards)
        grand_total = total_spare + total_labour + total_outsource

        vehicle_counts = Counter(job.get('vehicle_no') for job in job_cards if job.get('vehicle_no') and job.get('vehicle_no') not in {"-", "", "N/A"})
        top_vehicle, top_vehicle_count = (None, 0)
        if vehicle_counts:
            top_vehicle, top_vehicle_count = vehicle_counts.most_common(1)[0]

        site_counts = Counter(job.get('site') for job in job_cards if job.get('site'))
        top_site, top_site_count = (None, 0)
        if site_counts:
            top_site, top_site_count = site_counts.most_common(1)[0]

        average_spend = grand_total / total_jobs if total_jobs else 0.0

        metrics = [
            ("Total Job Cards", str(total_jobs)),
            ("Completed Jobs", str(completed_jobs)),
            ("In Progress Jobs", str(in_progress)),
            ("Unique Vehicles Serviced", str(len(unique_vehicles))),
            ("Unique Drivers", str(len(unique_drivers))),
            ("Active Sites", str(len(unique_sites))),
            ("Top Vehicle (jobs)", f"{top_vehicle} ({top_vehicle_count})" if top_vehicle else "—"),
            ("Top Site (jobs)", f"{top_site} ({top_site_count})" if top_site else "—"),
            ("Average Spend per Job", self._format_currency(average_spend)),
            ("Total Spare Parts Cost", self._format_currency(total_spare)),
            ("Total Labour Cost", self._format_currency(total_labour)),
            ("Total Outsource Cost", self._format_currency(total_outsource)),
            ("Grand Total Cost", self._format_currency(grand_total)),
        ]

        self.overview_table.setRowCount(len(metrics))
        for row, (metric, value) in enumerate(metrics):
            metric_item = QTableWidgetItem(metric)
            metric_item.setFlags(metric_item.flags() ^ Qt.ItemIsEditable)
            self.overview_table.setItem(row, 0, metric_item)
            value_item = QTableWidgetItem(value)
            value_item.setFont(QFont("Arial", 10, QFont.Bold))
            value_item.setFlags(value_item.flags() ^ Qt.ItemIsEditable)
            self.overview_table.setItem(row, 1, value_item)
        
        self.overview_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.overview_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
    
    def _load_engine_oil_usage(self, filters):
        """Aggregate most-used engine oils within the selected period."""
        usage = {}
        for job in self.filtered_jobcards:
            vehicle_no = job.get('vehicle_no')
            for part in job.get('spare_parts', []):
                description = part.get('description') or part.get('item_description') or part.get('category') or ""
                if not description:
                    continue
                lower_desc = description.lower()
                if "engine" not in lower_desc or "oil" not in lower_desc:
                    continue
                id_code = part.get('id_code') or description
                entry = usage.setdefault(id_code, {
                    "name": description,
                    "id_code": part.get('id_code'),
                    "quantity": 0.0,
                    "cost": 0.0,
                    "job_ids": set(),
                    "vehicles": set()
                })
                entry["quantity"] += self._to_float(part.get('quantity'))
                entry["cost"] += self._to_float(part.get('total'))
                entry["job_ids"].add(job['id'])
                if vehicle_no and vehicle_no not in {"", "-", "N/A"}:
                    entry["vehicles"].add(vehicle_no)

        sorted_usage = sorted(
            usage.values(),
            key=lambda data: (data["quantity"], data["cost"]),
            reverse=True
        )

        headers = ["Engine Oil", "Total Qty", "Jobs", "Vehicles", "Sample Vehicles", "Total Cost"]
        rows = []
        for record in sorted_usage[:50]:
            vehicles = sorted(record["vehicles"])
            rows.append({
                "values": [
                    record["name"],
                    f"{record['quantity']:.2f}",
                    str(len(record["job_ids"])),
                    str(len(record["vehicles"])),
                    ", ".join(vehicles[:3]) if vehicles else "—",
                    self._format_currency(record["cost"])
                ],
                "job_ids": list(record["job_ids"]),
                "context": {
                    "type": "part",
                    "id_codes": [record["id_code"]] if record["id_code"] else [],
                    "description": record["name"].lower()
                },
                "title": record["name"]
            })

        self._populate_table(self.engine_oil_table, headers, rows)

    def _load_vehicle_insights(self, filters):
        """Populate most-serviced vehicles and repeat repairs."""
        vehicle_stats = {}
        for job in self.filtered_jobcards:
            vehicle_no = job.get('vehicle_no')
            if not vehicle_no or vehicle_no in {"", "-", "N/A"}:
                continue

            stats = vehicle_stats.setdefault(vehicle_no, {
                "vehicle": vehicle_no,
                "make": job.get('make') or "—",
                "model": job.get('model') or "—",
                "jobs": 0,
                "first": job.get('start_date_dt'),
                "last": job.get('start_date_dt'),
                "spare": 0.0,
                "labour": 0.0,
                "outsource": 0.0,
                "job_ids": set()
            })

            stats["jobs"] += 1
            date_dt = job.get('start_date_dt')
            if date_dt:
                if not stats["first"] or date_dt < stats["first"]:
                    stats["first"] = date_dt
                if not stats["last"] or date_dt > stats["last"]:
                    stats["last"] = date_dt
            stats["spare"] += job.get('spare_cost', 0.0)
            stats["labour"] += job.get('labour_cost', 0.0)
            stats["outsource"] += job.get('outsource_cost', 0.0)
            stats["job_ids"].add(job['id'])

        sorted_stats = sorted(
            vehicle_stats.values(),
            key=lambda item: (item["jobs"], item["last"] or datetime.date.min),
            reverse=True
        )

        headers = ["Vehicle", "Make", "Model", "Jobs", "First Service", "Last Service", "Total Cost", "Spare", "Labour", "Outsource"]
        rows = []
        for entry in sorted_stats[:50]:
            total_cost = entry["spare"] + entry["labour"] + entry["outsource"]
            rows.append({
                "values": [
                    entry["vehicle"],
                    entry["make"],
                    entry["model"],
                    str(entry["jobs"]),
                    entry["first"].isoformat() if entry["first"] else "—",
                    entry["last"].isoformat() if entry["last"] else "—",
                    self._format_currency(total_cost),
                    self._format_currency(entry["spare"]),
                    self._format_currency(entry["labour"]),
                    self._format_currency(entry["outsource"])
                ],
                "job_ids": list(entry["job_ids"]),
                "context": {"type": "vehicle", "vehicle_no": entry["vehicle"]},
                "title": entry["vehicle"]
            })

        self._populate_table(self.vehicle_table, headers, rows)

        repeat_stats = {}
        for job in self.filtered_jobcards:
            vehicle_no = job.get('vehicle_no')
            section = job.get('section')
            if not vehicle_no or vehicle_no in {"", "-", "N/A"} or not section:
                continue
            key = (vehicle_no, section)
            data = repeat_stats.setdefault(key, {
                "vehicle": vehicle_no,
                "section": section,
                "jobs": 0,
                "last": job.get('start_date_dt'),
                "total": 0.0,
                "job_ids": set()
            })
            data["jobs"] += 1
            date_dt = job.get('start_date_dt')
            if date_dt and (not data["last"] or date_dt > data["last"]):
                data["last"] = date_dt
            data["total"] += job.get('total_cost', 0.0)
            data["job_ids"].add(job['id'])

        repeat_rows = []
        for data in repeat_stats.values():
            if data["jobs"] <= 1:
                continue
            repeat_rows.append({
                "values": [
                    data["vehicle"],
                    data["section"],
                    str(data["jobs"]),
                    data["last"].isoformat() if data["last"] else "—",
                    self._format_currency(data["total"])
                ],
                "job_ids": list(data["job_ids"]),
                "context": {"type": "repeat", "vehicle_no": data["vehicle"], "section": data["section"]},
                "title": f"{data['vehicle']} · {data['section']}"
            })

        repeat_rows.sort(key=lambda item: (int(item["values"][2]), item["values"][3]), reverse=True)
        self._populate_table(self.repeat_repairs_table, ["Vehicle", "Section", "Jobs", "Last Service", "Total Cost"], repeat_rows[:50])

    def _load_spare_part_trends(self, filters):
        """Highlight top spare parts during the trailing 7 days inside the selected window."""
        week_start_dt = max(
            filters['from_dt'],
            filters['to_dt'] - datetime.timedelta(days=6)
        )

        part_stats = {}
        for job in self.filtered_jobcards:
            date_dt = job.get('start_date_dt')
            if not date_dt or date_dt < week_start_dt:
                continue
            vehicle_no = job.get('vehicle_no')
            for part in job.get('spare_parts', []):
                description = part.get('description') or part.get('item_description') or ""
                if not description:
                    continue
                id_code = part.get('id_code') or description
                entry = part_stats.setdefault(id_code, {
                    "name": description,
                    "id_code": part.get('id_code'),
                    "quantity": 0.0,
                    "cost": 0.0,
                    "job_ids": set(),
                    "vehicles": set()
                })
                entry["quantity"] += self._to_float(part.get('quantity'))
                entry["cost"] += self._to_float(part.get('total'))
                entry["job_ids"].add(job['id'])
                if vehicle_no and vehicle_no not in {"", "-", "N/A"}:
                    entry["vehicles"].add(vehicle_no)

        sorted_parts = sorted(
            part_stats.values(),
            key=lambda data: (data["quantity"], data["cost"]),
            reverse=True
        )

        headers = ["Spare Part", "Jobs", "Vehicles", "Total Qty", "Total Cost", "Sample Vehicles"]
        rows = []
        for record in sorted_parts[:50]:
            vehicles = sorted(record["vehicles"])
            rows.append({
                "values": [
                    record["name"],
                    str(len(record["job_ids"])),
                    str(len(record["vehicles"])),
                    f"{record['quantity']:.2f}",
                    self._format_currency(record["cost"]),
                    ", ".join(vehicles[:3]) if vehicles else "—"
                ],
                "job_ids": list(record["job_ids"]),
                "context": {
                    "type": "part",
                    "id_codes": [record["id_code"]] if record["id_code"] else [],
                    "description": record["name"].lower()
                },
                "title": record["name"]
            })

        self._populate_table(self.spare_trends_table, headers, rows)

    def _load_cost_trends(self, filters):
        """Summarize cost breakdown grouped by service month."""
        trends = defaultdict(lambda: {"jobs": 0, "spare": 0.0, "labour": 0.0, "outsource": 0.0, "job_ids": set()})
        for job in self.filtered_jobcards:
            date_dt = job.get('start_date_dt')
            if not date_dt:
                continue
            period = date_dt.strftime("%Y-%m")
            entry = trends[period]
            entry["jobs"] += 1
            entry["spare"] += job.get('spare_cost', 0.0)
            entry["labour"] += job.get('labour_cost', 0.0)
            entry["outsource"] += job.get('outsource_cost', 0.0)
            entry["job_ids"].add(job['id'])

        rows = []
        for period, data in trends.items():
            total = data["spare"] + data["labour"] + data["outsource"]
            rows.append({
                "values": [
                    period,
                    str(data["jobs"]),
                    self._format_currency(data["spare"]),
                    self._format_currency(data["labour"]),
                    self._format_currency(data["outsource"]),
                    self._format_currency(total)
                ],
                "job_ids": list(data["job_ids"]),
                "context": {"type": "period", "period": period},
                "title": period
            })

        rows.sort(key=lambda row: row["values"][0], reverse=True)
        self._populate_table(self.cost_table, ["Period (YYYY-MM)", "Jobs", "Spare", "Labour", "Outsource", "Total"], rows)

    def _populate_table(self, table, headers, rows):
        """Populate helper ensuring metadata is attached for drilldowns."""
        table.clear()
        table.setColumnCount(len(headers))
        table.setHorizontalHeaderLabels(headers)
        table.setRowCount(len(rows))

        for row_idx, row in enumerate(rows):
            values = row.get("values", [])
            metadata = {
                "job_ids": row.get("job_ids", []),
                "context": row.get("context"),
                "title": row.get("title")
            }
            for col_idx, value in enumerate(values):
                display = "—" if value in (None, "") else value
                item = QTableWidgetItem(display)
                item.setFlags(item.flags() ^ Qt.ItemIsEditable)
                if col_idx == 0 and metadata["job_ids"]:
                    item.setData(Qt.UserRole, metadata)
                if col_idx > 0 and self._looks_numeric(display):
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                table.setItem(row_idx, col_idx, item)

        table.resizeColumnsToContents()
        if table.columnCount():
            table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)

    def _looks_numeric(self, value):
        """Heuristic to detect numeric strings for alignment."""
        if value is None:
            return False
        text = str(value).replace('Rs.', '').replace(',', '').replace('%', '').strip()
        if not text:
            return False
        try:
            float(text)
            return True
        except ValueError:
            return False

    def _register_drilldown(self, table):
        table.cellDoubleClicked.connect(lambda row, col, t=table: self._open_drilldown_from_table(t, row))

    def _open_drilldown_from_table(self, table, row):
        item = table.item(row, 0)
        if not item:
            return
        payload = item.data(Qt.UserRole)
        if not payload:
            return
        job_ids = payload.get('job_ids') or []
        if not job_ids:
            return
        title = payload.get('title') or item.text()
        self._show_drilldown_dialog(title, job_ids, payload.get('context'))

    def _show_drilldown_dialog(self, title, job_ids, context):
        jobs = [self.filtered_job_index.get(job_id) for job_id in job_ids if self.filtered_job_index.get(job_id)]
        if not jobs:
            QMessageBox.information(self, "No Data", "No job details were found for this selection.")
            return

        dialog = QDialog(self)
        dialog.setWindowTitle(f"Drill-down · {title}")
        dialog.setMinimumWidth(780)
        layout = QVBoxLayout(dialog)

        info_label = QLabel(f"{len(jobs)} job(s) matched the selected analytics row.")
        info_label.setStyleSheet("color: #4b5563; font-size: 11px; font-weight: 500;")
        layout.addWidget(info_label)

        job_headers = ["Job No", "Vehicle", "Site", "Section", "Start", "End", "Status", "Total Cost"]
        job_table = QTableWidget()
        job_table.setAlternatingRowColors(True)
        job_table.setEditTriggers(QTableWidget.NoEditTriggers)
        job_table.setSelectionBehavior(QTableWidget.SelectRows)
        job_table.setColumnCount(len(job_headers))
        job_table.setHorizontalHeaderLabels(job_headers)
        job_table.setRowCount(len(jobs))

        for row_idx, job in enumerate(jobs):
            values = [
                job.get('job_no') or "—",
                job.get('vehicle_no') or "—",
                job.get('site') or "—",
                job.get('section') or "—",
                job.get('start_date') or "—",
                job.get('end_date') or "—",
                job.get('status') or "—",
                self._format_currency(job.get('total_cost', 0.0))
            ]
            for col_idx, value in enumerate(values):
                item = QTableWidgetItem(value)
                item.setFlags(item.flags() ^ Qt.ItemIsEditable)
                if col_idx == 7 and self._looks_numeric(value):
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                job_table.setItem(row_idx, col_idx, item)

        job_table.resizeColumnsToContents()
        layout.addWidget(job_table)

        items = []
        for job in jobs:
            items.extend(self._gather_items_for_context(job, context))

        if items:
            item_headers = ["Category", "Description", "Qty/Hrs", "Cost", "Job No", "Vehicle"]
            item_table = QTableWidget()
            item_table.setAlternatingRowColors(True)
            item_table.setEditTriggers(QTableWidget.NoEditTriggers)
            item_table.setSelectionBehavior(QTableWidget.SelectRows)
            item_table.setColumnCount(len(item_headers))
            item_table.setHorizontalHeaderLabels(item_headers)
            item_table.setRowCount(len(items))

            for row_idx, item_data in enumerate(items):
                values = [
                    item_data.get('category', '—'),
                    item_data.get('description', '—'),
                    item_data.get('quantity', '—'),
                    item_data.get('cost', '—'),
                    item_data.get('job_no', '—'),
                    item_data.get('vehicle', '—')
                ]
                for col_idx, value in enumerate(values):
                    item_widget = QTableWidgetItem(value)
                    item_widget.setFlags(item_widget.flags() ^ Qt.ItemIsEditable)
                    if col_idx in {2, 3} and self._looks_numeric(value):
                        item_widget.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    item_table.setItem(row_idx, col_idx, item_widget)

            item_table.resizeColumnsToContents()
            layout.addWidget(item_table)
        else:
            empty_label = QLabel("No spare parts, labour, or outsource entries matched this drill-down filter.")
            empty_label.setStyleSheet("color: #6b7280; font-size: 11px;")
            layout.addWidget(empty_label)

        buttons = QDialogButtonBox(QDialogButtonBox.Close)
        buttons.rejected.connect(dialog.reject)
        buttons.accepted.connect(dialog.accept)
        layout.addWidget(buttons)

        dialog.exec()

    def _gather_items_for_context(self, job, context):
        """Collect spare, labour, and outsource items filtered by context."""
        ctx_type = context.get('type') if context else None
        id_codes = [code for code in (context.get('id_codes') or []) if code]
        keyword = (context.get('description') or "").lower() if context else ""
        items = []

        def append_item(category, description, quantity, cost):
            items.append({
                'category': category,
                'description': description,
                'quantity': quantity if quantity not in (None, '') else '—',
                'cost': self._format_currency(cost),
                'job_no': job.get('job_no') or '—',
                'vehicle': job.get('vehicle_no') or '—'
            })

        for part in job.get('spare_parts', []):
            description = part.get('description') or part.get('item_description') or ""
            part_id = part.get('id_code')
            lower_desc = description.lower()
            if ctx_type == 'part':
                if id_codes and part_id not in id_codes:
                    continue
                if not id_codes and keyword and keyword not in lower_desc:
                    continue
            elif ctx_type in {'vehicle', 'repeat', 'period', None}:
                pass
            else:
                continue
            label = f"{part_id or ''} {description}".strip()
            append_item('Spare Part', label, part.get('quantity') or part.get('qty'), self._to_float(part.get('total')))

        if ctx_type in {'vehicle', 'repeat', 'period', None}:
            for work in job.get('labour_works', []):
                desc = work.get('labour_name') or work.get('description') or work.get('grade') or 'Labour'
                append_item('Labour', desc, work.get('hours') or work.get('work_hours') or work.get('duration'), self._to_float(work.get('work_cost')))

            for work in job.get('outsource_works', []):
                desc = work.get('work_type') or work.get('description') or 'Outsource'
                # Use work date in quantity slot to convey timing if no numeric value
                qty = work.get('work_date') or work.get('remark')
                append_item('Outsource', desc, qty, self._to_float(work.get('cost')))

        return items

    def _format_currency(self, value):
        try:
            return f"Rs. {float(value):,.2f}"
        except (TypeError, ValueError):
            return "Rs. 0.00"

    def _to_float(self, value):
        if value in (None, "", "-"):
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        try:
            return float(str(value).replace(",", "").strip())
        except ValueError:
            return 0.0

    def _resolve_active_table_for_export(self):
        index = self.tabs.currentIndex()
        if index == 0:
            return self.overview_table, "Overview"
        if index == 1:
            return self.engine_oil_table, "Engine_Oils"
        if index == 2:
            table = self.repeat_repairs_table if self.repeat_repairs_table.hasFocus() else self.vehicle_table
            name = "Repeat_Repairs" if table is self.repeat_repairs_table else "Vehicle_Insights"
            return table, name
        if index == 3:
            return self.spare_trends_table, "Spare_Part_Trends"
        if index == 4:
            return self.cost_table, "Cost_Trends"
        return None, "Report"

    def _ensure_indexes(self):
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            statements = [
                "CREATE INDEX IF NOT EXISTS idx_job_cards_start_date ON job_cards(start_date)",
                "CREATE INDEX IF NOT EXISTS idx_job_cards_site ON job_cards(site)",
                "CREATE INDEX IF NOT EXISTS idx_job_cards_status ON job_cards(status)",
                "CREATE INDEX IF NOT EXISTS idx_job_cards_vehicle ON job_cards(vehicle_no)",
                "CREATE INDEX IF NOT EXISTS idx_job_cards_section ON job_cards(section)"
            ]
            for sql in statements:
                cursor.execute(sql)
            conn.commit()
        except sqlite3.Error:
            pass
        finally:
            try:
                conn.close()
            except Exception:
                pass

    def _build_filter_clause(self, alias="jc"):
        from_str = self.from_date.date().toString("yyyy-MM-dd")
        to_str = self.to_date.date().toString("yyyy-MM-dd")
        conditions = [f"{alias}.start_date >= ?", f"{alias}.start_date <= ?"]
        params = [from_str, to_str]

        if self.site_filter.currentText() != "All Sites":
            conditions.append(f"{alias}.site = ?")
            params.append(self.site_filter.currentText())

        if self.status_filter.currentText() != "All Status":
            conditions.append(f"{alias}.status = ?")
            params.append(self.status_filter.currentText())

        clause = f"WHERE {' AND '.join(conditions)}"
        from_dt = datetime.date.fromisoformat(from_str)
        to_dt = datetime.date.fromisoformat(to_str)
        return {
            'clause': clause,
            'params': params,
            'from_date': from_str,
            'to_date': to_str,
            'from_dt': from_dt,
            'to_dt': to_dt
        }

    def _get_filtered_jobcards(self, filters):
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        query = f"""
            SELECT id, job_no, company_no, vehicle_no, driver, make, model, type,
                   site, section, start_date, end_date, status,
                   spare_parts, labour_works, outsource_works, description
            FROM job_cards jc
            {filters['clause']}
            ORDER BY start_date DESC, id DESC
        """
        cursor.execute(query, filters['params'])
        rows = cursor.fetchall()
        conn.close()

        jobcards = []
        for row in rows:
            start_date = row['start_date'] or ""
            try:
                start_dt = datetime.date.fromisoformat(start_date) if start_date else None
            except ValueError:
                start_dt = None

            spare_parts = self._safe_json_loads(row['spare_parts'], [])
            labour_works = self._safe_json_loads(row['labour_works'], [])
            outsource_works = self._safe_json_loads(row['outsource_works'], [])

            spare_cost = sum(self._to_float(part.get('total')) for part in spare_parts)
            labour_cost = sum(self._to_float(work.get('work_cost')) for work in labour_works)
            outsource_cost = sum(self._to_float(work.get('cost')) for work in outsource_works)

            jobcards.append({
                'id': row['id'],
                'job_no': row['job_no'],
                'company_no': row['company_no'],
                'vehicle_no': row['vehicle_no'],
                'driver': row['driver'],
                'make': row['make'],
                'model': row['model'],
                'type': row['type'],
                'site': row['site'],
                'section': row['section'],
                'start_date': start_date,
                'start_date_dt': start_dt,
                'end_date': row['end_date'] or "",
                'status': row['status'],
                'description': row['description'],
                'spare_parts': spare_parts,
                'labour_works': labour_works,
                'outsource_works': outsource_works,
                'spare_cost': spare_cost,
                'labour_cost': labour_cost,
                'outsource_cost': outsource_cost,
                'total_cost': spare_cost + labour_cost + outsource_cost
            })

        self.filtered_job_index = {job['id']: job for job in jobcards}
        return jobcards

    def _safe_json_loads(self, payload, default=None):
        if not payload:
            return [] if default is None else list(default) if isinstance(default, list) else default
        try:
            data = json.loads(payload)
            if isinstance(data, list):
                return data
        except (TypeError, ValueError, json.JSONDecodeError):
            pass
        return [] if default is None else list(default) if isinstance(default, list) else default
    
    def export_csv(self):
        """Export current tab data to CSV"""
        current_index = self.tabs.currentIndex()
        tables = [
            self.overview_table, self.spare_parts_table, 
            self.vehicle_table, self.cost_table, self.performance_table
        ]
        tab_names = ["Overview", "Spare_Parts", "Vehicle_Analysis", "Cost_Analysis", "Performance"]
        
        table = tables[current_index]
        tab_name = tab_names[current_index]
        
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Export to CSV",
            f"Report_{tab_name}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            "CSV Files (*.csv)"
        )
        
        if not filename:
            return
        
        try:
            with open(filename, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                
                # Write headers
                headers = [table.horizontalHeaderItem(i).text() for i in range(table.columnCount())]
                writer.writerow(headers)
                
                # Write data
                for row in range(table.rowCount()):
                    row_data = []
                    for col in range(table.columnCount()):
                        item = table.item(row, col)
                        row_data.append(item.text() if item else "")
                    writer.writerow(row_data)
            
            QMessageBox.information(self, "Success", f"Report exported successfully to:\n{filename}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to export CSV:\n{str(e)}")
    
    def export_xlsx(self):
        """Export current tab data to Excel"""
        if not HAS_OPENPYXL:
            QMessageBox.warning(self, "Not Available", "Excel export requires openpyxl library.\nInstall with: pip install openpyxl")
            return
        
        current_index = self.tabs.currentIndex()
        tables = [
            self.overview_table, self.spare_parts_table,
            self.vehicle_table, self.cost_table, self.performance_table
        ]
        tab_names = ["Overview", "Spare_Parts", "Vehicle_Analysis", "Cost_Analysis", "Performance"]
        
        table = tables[current_index]
        tab_name = tab_names[current_index]
        
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Export to Excel",
            f"Report_{tab_name}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel Files (*.xlsx)"
        )
        
        if not filename:
            return
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = tab_name
            
            # Styling
            header_fill = PatternFill(start_color="2E7D6E", end_color="2E7D6E", fill_type="solid")
            header_font = XLFont(color="FFFFFF", bold=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Write headers
            for col in range(table.columnCount()):
                cell = ws.cell(row=1, column=col+1)
                cell.value = table.horizontalHeaderItem(col).text()
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Write data
            for row in range(table.rowCount()):
                for col in range(table.columnCount()):
                    item = table.item(row, col)
                    cell = ws.cell(row=row+2, column=col+1)
                    cell.value = item.text() if item else ""
                    cell.border = border
                    
                    # Right-align numbers
                    if item and ("Rs." in item.text() or "%" in item.text() or item.text().replace(".", "").replace(",", "").isdigit()):
                        cell.alignment = Alignment(horizontal='right')
            
            # Auto-fit columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
            
            wb.save(filename)
            QMessageBox.information(self, "Success", f"Report exported successfully to:\n{filename}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to export Excel:\n{str(e)}")
    
    def export_pdf(self):
        """Export current tab data to PDF"""
        if not HAS_REPORTLAB:
            QMessageBox.warning(self, "Not Available", "PDF export requires reportlab library.\nInstall with: pip install reportlab")
            return
        
        current_index = self.tabs.currentIndex()
        tables = [
            self.overview_table, self.spare_parts_table,
            self.vehicle_table, self.cost_table, self.performance_table
        ]
        tab_names = ["Overview", "Spare Parts Analysis", "Vehicle Analysis", "Cost Analysis", "Performance Metrics"]
        
        table = tables[current_index]
        tab_name = tab_names[current_index]
        
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Export to PDF",
            f"Report_{tab_name.replace(' ', '_')}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
            "PDF Files (*.pdf)"
        )
        
        if not filename:
            return
        
        try:
            doc = SimpleDocTemplate(filename, pagesize=A4, topMargin=0.75*inch, bottomMargin=0.75*inch)
            story = []
            styles = getSampleStyleSheet()
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor("#2E7D6E"),
                spaceAfter=12,
                alignment=TA_CENTER
            )
            story.append(Paragraph(f"Analytics Report: {tab_name}", title_style))
            story.append(Paragraph(
                f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", 
                styles['Normal']
            ))
            story.append(Spacer(1, 0.3*inch))
            
            # Table data
            data = []
            headers = [table.horizontalHeaderItem(i).text() for i in range(table.columnCount())]
            data.append(headers)
            
            for row in range(table.rowCount()):
                row_data = []
                for col in range(table.columnCount()):
                    item = table.item(row, col)
                    row_data.append(item.text() if item else "")
                data.append(row_data)
            
            # Create PDF table
            pdf_table = Table(data)
            pdf_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2E7D6E")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
            ]))
            
            story.append(pdf_table)
            
            # Build PDF
            doc.build(story)
            QMessageBox.information(self, "Success", f"Report exported successfully to:\n{filename}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to export PDF:\n{str(e)}")
    
    def go_back(self):
        """Navigate back to home page"""
        if self.parent:
            self.parent.go_to_home()
